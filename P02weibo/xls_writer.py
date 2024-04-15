import os

from openpyxl import load_workbook, Workbook


class XlsWriter(object):
    _entered = False

    def __init__(self, file_name, headers=None):
        self.file_name = file_name
        if os.path.exists(file_name):
            self.workbook = load_workbook(file_name)
            self.sheet = self.workbook.active
        else:
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            if headers:
                self.set_headers(headers)
        self.headers = headers if headers else []
        self.column_count = len(self.headers)

    def __enter__(self):
        self._entered = True  # 标记已进入上下文管理器
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.workbook.save(self.file_name)
        self._entered = False  # 重置标志
        if exc_type:
            raise

    def _ensure_entered(self):
        """确保已通过上下文管理器进入实例"""
        if not self._entered:
            raise RuntimeError("This instance must be used as a context manager.")

    def set_headers(self, headers):
        if self.sheet.max_row == 1:  # Only write headers if the sheet is empty
            for col_num, header in enumerate(headers):
                self.sheet.cell(row=1, column=col_num + 1, value=header)
        self.column_count = len(headers)

    def add_item(self, item):
        self._ensure_entered()  # 添加检查
        if len(item) != self.column_count:
            raise ValueError("Item length does not match the number of columns.")
        row_num = self.sheet.max_row + 1
        for col_num, value in enumerate(item):
            self.sheet.cell(row=row_num, column=col_num + 1, value=value)


# Usage example
if __name__ == "__main__":
    with XlsWriter("example.xlsx", headers=["Name", "Age"]) as writer:
        writer.add_item(["Alice", 30])
        writer.add_item(["Bob", 25])
